"""stocks.json 저장 위치, 자동 마이그레이션, Excel import/export."""

import os
import sys
import shutil
from pathlib import Path

from .portfolio import portfolio_totals


# ─── 설정 파일 경로 (OS별 표준 디렉토리) ──────────────────────────────────────
def _config_dir() -> Path:
    if sys.platform == "win32":
        base = Path(os.environ.get("APPDATA") or (Path.home() / "AppData" / "Roaming"))
    elif sys.platform == "darwin":
        base = Path.home() / "Library" / "Application Support"
    else:
        base = Path(os.environ.get("XDG_CONFIG_HOME") or (Path.home() / ".config"))
    d = base / "Pinstock"
    d.mkdir(parents=True, exist_ok=True)
    return d


CONFIG_FILE = str(_config_dir() / "stocks.json")
BACKUP_FILE = CONFIG_FILE + ".bak"


# ─── 종목 스키마 기본값 ─────────────────────────────────────────────────────
MARKET_KR = "KR"
MARKET_US = "US"
CURRENCY_KRW = "KRW"
CURRENCY_USD = "USD"


def normalize_stock_schema(stock: dict) -> dict:
    """기존 stocks.json 항목에 시장/통화 기본값을 보강한다.

    미국 주식 지원 전의 기존 데이터는 market/currency 필드가 없으므로 한국
    주식으로 취급한다. 알 수 없는 부가 필드(pos, hidden 등)는 그대로 보존한다.
    """
    normalized = dict(stock)
    market = str(normalized.get("market") or MARKET_KR).strip().upper()
    if market not in {MARKET_KR, MARKET_US}:
        market = MARKET_KR
    normalized["market"] = market

    default_currency = CURRENCY_USD if market == MARKET_US else CURRENCY_KRW
    currency = str(normalized.get("currency") or default_currency).strip().upper()
    normalized["currency"] = currency or default_currency

    if market == MARKET_US and "buy_exchange_rate" in normalized:
        try:
            normalized["buy_exchange_rate"] = float(normalized["buy_exchange_rate"])
        except (TypeError, ValueError):
            normalized.pop("buy_exchange_rate", None)

    return normalized


def normalize_stocks_schema(stocks: list[dict]) -> list[dict]:
    return [normalize_stock_schema(s) for s in stocks if isinstance(s, dict)]


# ─── 레거시 위치(레포 루트/CWD)에서 새 위치로 1회 자동 이전 ───────────────
def migrate_legacy_config() -> None:
    """저장소 루트(또는 현재 작업 디렉토리)에 있던 stocks.json 을 새 위치로 1회 이전.

    v1.x 시절 단일 스크립트 옆에 stocks.json 을 저장하던 기존 사용자를 위함.
    새 위치에 이미 파일이 있으면 아무것도 하지 않는다.
    이전 후 옛 파일은 같은 경로에 `.migrated` 마커를 남겨 확인할 수 있게 한다.
    """
    if os.path.exists(CONFIG_FILE):
        return
    candidates = [
        # 레포 루트 (pinstock/core/storage.py 의 두 단계 부모)
        Path(__file__).resolve().parent.parent.parent / "stocks.json",
        # 현재 작업 디렉토리
        Path.cwd() / "stocks.json",
    ]
    seen: set[Path] = set()
    for legacy in candidates:
        try:
            legacy = legacy.resolve()
        except Exception:
            continue
        if legacy in seen:
            continue
        seen.add(legacy)
        if legacy.is_file():
            try:
                shutil.move(str(legacy), CONFIG_FILE)
                marker = Path(str(legacy) + ".migrated")
                marker.write_text(
                    f"이 파일은 새 위치로 이동되었습니다:\n{CONFIG_FILE}\n",
                    encoding="utf-8",
                )
                print(f"[migrate] stocks.json 을 {CONFIG_FILE} 로 이전했습니다.")
            except Exception as e:
                print(f"[migrate] 오류: {e}")
            return


# ─── Excel import/export 컬럼 정의 ────────────────────────────────────────────
# 헤더 ↔ stocks.json 필드 매핑. 순서는 export 시 컬럼 순서가 됨.
EXCEL_COLUMNS = [
    ("종목코드", "code"),
    ("종목명",   "name"),
    ("평단가",   "avg_price"),
    ("수량",     "quantity"),
]


# ─── Excel import/export ─────────────────────────────────────────────────────
def export_stocks_to_excel(stocks: list[dict], path: str,
                           current_prices: dict | None = None,
                           usd_krw_rate: float | None = None) -> None:
    """보유 종목을 .xlsx 로 내보내기.
    - 종목코드는 텍스트 셀로 저장 (선행 0 보존: '005930', '0183J0').
    - 위젯 위치(pos)는 제외 — 다른 PC에서는 화면 좌표가 달라 의미가 없음.
    - current_prices ({code: price}) 와 usd_krw_rate 가 주어지면 시트 하단에
      포트폴리오 요약(총 매입금액 / 평가금액 / 평가손익 / 수익률)을 빈 행 한
      줄로 분리해서 추가. 미국 주식은 원화 기준으로 합산한다.
      import 시에는 빈 행 이후의 행을 모두 무시하므로 라운드트립에 영향 없음."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "보유종목"

    headers = [h for h, _ in EXCEL_COLUMNS]
    ws.append(headers)
    bold = Font(bold=True)
    for col_idx in range(1, len(headers) + 1):
        ws.cell(row=1, column=col_idx).font = bold

    for s in stocks:
        row = []
        for _, key in EXCEL_COLUMNS:
            if key == "code":
                row.append(str(s.get("code", "")))
            elif key == "name":
                row.append(s.get("name", s.get("code", "")))
            elif key == "quantity":
                row.append(float(s.get(key, 0)))
            else:
                row.append(int(s.get(key, 0)))
        ws.append(row)

    # 종목코드 컬럼을 텍스트 포맷으로 (선행 0/영문 안전)
    code_col_idx = next(i for i, (_, k) in enumerate(EXCEL_COLUMNS, 1) if k == "code")
    code_letter = ws.cell(row=1, column=code_col_idx).column_letter
    for cell in ws[code_letter][1:]:   # 헤더 제외
        cell.number_format = "@"
        cell.alignment = Alignment(horizontal="left")

    # 컬럼 너비 자동 조정 (간단히 헤더+여유)
    widths = {"종목코드": 12, "종목명": 28, "평단가": 12, "수량": 10}
    for col_idx, (header, _) in enumerate(EXCEL_COLUMNS, 1):
        letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[letter].width = widths.get(header, 14)

    # ── 포트폴리오 요약 (종목이 1개 이상일 때) ────────────────────────
    # 종목 표와 빈 행 한 줄로 분리. import 측에서 빈 행 이후를 모두 무시하므로
    # 라운드트립 안전.
    if stocks:
        totals = portfolio_totals(
            stocks,
            current_prices=current_prices,
            usd_krw_rate=usd_krw_rate,
            include_hidden=True,
        )
        total_invest = totals["total_invest"]
        total_eval = totals["total_eval"]
        profit = totals["profit"]
        prate = totals["profit_rate"]

        # 빈 행 한 줄 띄우고 다음 행에 요약 헤더
        header_row = ws.max_row + 2
        ws.cell(row=header_row, column=1, value="포트폴리오 요약").font = bold

        rows = [
            ("총 매입금액", total_invest, "#,##0"),
            ("평가금액",   total_eval,   "#,##0"),
            ("평가손익",   profit,        "#,##0"),
            ("수익률 (%)", round(prate, 2), "0.00"),
        ]
        for i, (label, val, fmt) in enumerate(rows, 1):
            r = header_row + i
            ws.cell(row=r, column=1, value=label)
            val_cell = ws.cell(row=r, column=2, value=val)
            val_cell.number_format = fmt
            val_cell.alignment = Alignment(horizontal="right")

    wb.save(path)


def import_stocks_from_excel(path: str) -> list[dict]:
    """Excel 파일에서 보유 종목을 읽어 stocks.json 형식 dict 리스트로 반환.
    검증 실패 시 ValueError 를 발생시킨다 (메시지는 사용자에게 그대로 표시 가능)."""
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    ws = wb.active
    if ws.max_row < 1:
        raise ValueError("시트가 비어 있습니다.")

    # 1행 헤더 읽기 (공백/None 안전)
    header_row = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    required = [h for h, _ in EXCEL_COLUMNS]
    missing = [h for h in required if h not in header_row]
    if missing:
        raise ValueError(
            "필수 컬럼이 누락되었습니다: " + ", ".join(missing)
            + f"\n(필요한 헤더: {', '.join(required)})"
        )

    # 헤더명 → 컬럼 인덱스
    idx_of = {h: header_row.index(h) for h in required}

    stocks: list[dict] = []
    seen_codes: set[str] = set()
    errors: list[str] = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # 빈 행을 만나면 그 이후는 모두 무시 (export 시 빈 행으로 구분한 요약 섹션 등)
        if row is None or all(v is None or (isinstance(v, str) and not v.strip()) for v in row):
            break

        def cell(h: str):
            i = idx_of[h]
            return row[i] if i < len(row) else None

        raw_code = cell("종목코드")
        raw_name = cell("종목명")
        raw_avg  = cell("평단가")
        raw_qty  = cell("수량")

        # 종목코드: 숫자로 읽혔어도 문자열로 정규화 후 6자 영숫자 검증 + 대문자
        if raw_code is None or str(raw_code).strip() == "":
            errors.append(f"{row_num}행: 종목코드가 비어 있습니다.")
            continue
        code = str(raw_code).strip().upper()
        # 엑셀이 숫자로 인식해 선행 0 손실된 경우 6자리로 패딩 (전부 숫자일 때만)
        if code.isdigit() and len(code) < 6:
            code = code.zfill(6)
        if len(code) != 6 or not code.isalnum():
            errors.append(f"{row_num}행: 종목코드 '{code}' 가 6자리 영숫자가 아닙니다.")
            continue
        if code in seen_codes:
            errors.append(f"{row_num}행: 종목코드 '{code}' 가 중복되었습니다.")
            continue

        # 평단가/수량 변환. 수량은 소수점 3자리까지 허용한다.
        try:
            avg_price = int(float(raw_avg)) if raw_avg is not None and str(raw_avg).strip() != "" else 0
        except (TypeError, ValueError):
            errors.append(f"{row_num}행: 평단가 '{raw_avg}' 가 숫자가 아닙니다.")
            continue
        try:
            quantity = round(float(raw_qty), 3) if raw_qty is not None and str(raw_qty).strip() != "" else 0
        except (TypeError, ValueError):
            errors.append(f"{row_num}행: 수량 '{raw_qty}' 가 숫자가 아닙니다.")
            continue
        if avg_price < 1:
            errors.append(f"{row_num}행: 평단가가 1 이상이어야 합니다.")
            continue
        if quantity <= 0:
            errors.append(f"{row_num}행: 수량이 0보다 커야 합니다.")
            continue

        name = str(raw_name).strip() if raw_name is not None and str(raw_name).strip() else code

        stocks.append(normalize_stock_schema({
            "code":      code,
            "name":      name,
            "avg_price": avg_price,
            "quantity":  quantity,
        }))
        seen_codes.add(code)

    if errors:
        # 너무 길지 않게 상위 10개만 보여줌
        head = "\n".join(errors[:10])
        more = f"\n... 외 {len(errors) - 10}건" if len(errors) > 10 else ""
        raise ValueError("다음 항목에서 오류가 발생했습니다:\n\n" + head + more)

    if not stocks:
        raise ValueError("가져올 종목이 없습니다. (데이터 행을 찾지 못했습니다)")

    return normalize_stocks_schema(stocks)
